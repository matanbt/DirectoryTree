import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter

# curr path by default, could alternatively be a given path:
PATH = os.getcwd()

# name of the folder being inspected
CURR_DIR_NAME = PATH.split("\\")[-1]


def pathToDict(path):
    """ given directory returns a python dict:
    key - relative path of inner dir, value - list of sunfolders and files containd directly by the dir """

    dir_dict = {}

    # 'depth' of the given path
    curr_path_len = len(path.split("\\")) - 1

    for currPath, subFolders, files in os.walk(path):
        dir_key = "\\".join(currPath.split("\\")[curr_path_len:])  # builds relative path
        dir_val = files + subFolders  # without marking subfolders
        dir_dict[dir_key] = dir_val

    return dir_dict


def rec_build_str(dir_dict, path, depth=0, main_key="-"):
    """builds directory hirerarchy string fits the given directory"""

    out_str = (main_key + " ") * depth + path.split("\\")[-1] + "\n"

    if path not in dir_dict:
        return out_str

    for obj in dir_dict[path]:
        out_str += rec_build_str(dir_dict, path + "\\" + obj, depth + 1)

    return out_str


def rec_build_depth(dir_dict, path, depth=0):
    """builds directory hirerarchy list fits the given directory
    each tuple represents: (depth, file=-1 \ folder=size of its [decendents includes], name) """

    if path not in dir_dict:
        return [(depth, -1, path.split("\\")[-1])]

    lst = [None]  # first element will be added later

    for obj in dir_dict[path]:
        lst += rec_build_depth(dir_dict, path + "\\" + obj, depth + 1)

    count_files = sum([1 for elem in lst[1:] if elem[1] == -1 or elem[1] == 0])  # counts files and empty folders
    lst[0] = (depth, count_files, path.split("\\")[-1])
    return lst


#excel:
def styleExcelInit(ws, r1, rn, c1, cn):
    # col width:
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 18
    # cells style:
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                    outline=Side(border_style=None, color='000000'),
                    vertical=Side(border_style=None, color='000000'),
                    horizontal=Side(border_style=None, color='000000'))
    alignment = Alignment(horizontal='center', vertical='center',
                          text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)

    rows = ws.iter_rows(r1, rn, c1, cn)
    for row in rows:
        for cell in row:
            cell.border = border
            cell.alignment = alignment


def styleFolderCell(ws, curr_col, curr_row):
    ws[curr_col + str(curr_row)].fill = PatternFill("solid", fgColor="DDDDDD")
    ws[curr_col + str(curr_row)].font = Font(b=True)


def exportToExcel(path):
    dir_dict = pathToDict(path)
    curr_dir_name = path.split("\\")[-1]
    final_lst = rec_build_depth(dir_dict, curr_dir_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dir Analysis"
    curr_row = 1
    col = 'A'

    styleExcelInit(ws, 1, final_lst[0][1], 1, max([i[0] for i in final_lst]) + 1)

    for obj in final_lst:

        # writes to cell
        curr_col = chr(ord(col) + obj[0])
        ws[curr_col + str(curr_row)] = obj[2]

        if obj[1] >= 0:  # it's a folder, could be empty
            styleFolderCell(ws, curr_col, curr_row)

        if obj[1] > 0:  # it's a non-empty-folder with files
            ws.merge_cells(curr_col + str(curr_row) + ":" + curr_col + str(curr_row + obj[1] - 1))
        else:
            curr_row += 1

    wb.save(curr_dir_name + "-Analysis.xlsx")
    print("'" + curr_dir_name + "-Analysis.xlsx' was created \n")

#TXT:
def exportToTXT(path):
    dir_dict = pathToDict(path)
    curr_dir_name = path.split("\\")[-1]
    with open(curr_dir_name + "-Analysis.txt", 'x') as f:
        f.write(rec_build_str(dir_dict, curr_dir_name))

#action:
exportToExcel(PATH)
exportToTXT(PATH)

